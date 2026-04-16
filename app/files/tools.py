from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.sandbox import require_sandbox_context


def _resolve_upload_path(file_path: str) -> Path:
    raw = (file_path or "").strip()
    if not raw:
        raise ValueError("file_path is required")
    if ".." in Path(raw).parts:
        raise ValueError("file_path cannot contain '..'")

    context = require_sandbox_context()
    relative = raw[1:] if raw.startswith("/") else raw
    target = (context.workspace_path / relative).resolve()
    root = context.workspace_path.resolve()
    if root not in target.parents and target != root:
        raise ValueError("file_path must stay inside the current sandbox workspace")
    if not target.exists() or not target.is_file():
        raise FileNotFoundError(f"file not found: {raw}")
    return target


def _cell_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def _parse_xlsx(path: Path, *, max_rows_per_sheet: int) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets: list[dict[str, Any]] = []
    try:
        for sheet in wb.worksheets:
            rows: list[list[str]] = []
            for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if idx > max_rows_per_sheet:
                    break
                rows.append([_cell_value(value) for value in row])
            sheets.append(
                {
                    "name": sheet.title,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "preview_rows": rows,
                    "truncated": (sheet.max_row or 0) > max_rows_per_sheet,
                }
            )
    finally:
        wb.close()
    return {"type": "xlsx", "sheets": sheets}


async def parse_excel_file(file_path: str, max_rows_per_sheet: int = 50) -> str:
    """
    Parse an uploaded Excel file from the current session sandbox.

    Args:
        file_path: Sandbox virtual path, for example /uploads/demo.xlsx.
        max_rows_per_sheet: Preview rows per sheet, default 50.
    """
    path = _resolve_upload_path(file_path)
    max_rows_per_sheet = max(1, min(int(max_rows_per_sheet or 50), 200))
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        result = _parse_xlsx(path, max_rows_per_sheet=max_rows_per_sheet)
    else:
        raise ValueError("Only .xlsx and .xlsm Excel files are supported.")

    result.update(
        {
            "ok": True,
            "file_path": file_path,
            "filename": path.name,
            "size": path.stat().st_size,
        }
    )
    return json.dumps(result, ensure_ascii=False, default=str)


async def list_uploaded_excel_files() -> str:
    """List uploaded Excel files in the current session sandbox."""
    context = require_sandbox_context()
    upload_dir = context.workspace_path / "uploads"
    upload_dir.mkdir(parents=True, exist_ok=True)
    items = []
    for path in sorted(upload_dir.iterdir()):
        if path.is_file() and path.suffix.lower() in {".xlsx", ".xlsm"}:
            items.append({"file_path": f"/uploads/{path.name}", "filename": path.name, "size": path.stat().st_size})
    return json.dumps({"ok": True, "items": items}, ensure_ascii=False)
